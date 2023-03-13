import openpyxl
import os
import time
import xlrd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import re
from create_hawaii_alaska_xlsx_file import create_xlsx_alaska_hawaii


def read_file(file=None):
    zip_ranges = []

    # Открываем файл
    workbook = openpyxl.load_workbook(file or './Carrierszoneranges.xlsx')

    # Получаем список листов в файле
    sheet_names = workbook.sheetnames

    # Выбираем первый лист
    sheet = workbook[sheet_names[1]]

    # Получаем индекс колонки по названию
    column_index = None
    for cell in sheet[1]:
        if cell.value == 'UPS zone ranges':
            column_index = cell.column_letter
            break

    # Если колонка найдена, то выводим ее значения
    if column_index:
        column = sheet[column_index]
        for cell in column[1:]:
            if cell.value and 'CONCATENATE' in cell.value:
                val = re.findall(r'\w\d', cell.value)
                zip_ranges.append('-'.join([sheet[i].value for i in val]))
            else:
                zip_ranges.append(cell.value)

    # Закрываем файл
    workbook.close()

    return zip_ranges


def update_workbook(data, file=None):
    # открываем файл
    workbook = openpyxl.load_workbook(file or "Carrierszoneranges.xlsx")

    # Создаем новый лист
    sheet = workbook.create_sheet("Ups zip ranges(upd)")

    # определяем хедер
    headers = ["UPS zone ranges", "Zip From", "Zip To"]

    # добавляем хедер
    sheet.append(headers)

    for i in data:
        if i:
            zip_from, zip_to = i.split("-")
            row = [i, zip_from, zip_to]
            sheet.append(row)

    # сохраняем воркбук
    workbook.save(file or "Carrierszoneranges.xlsx")


def parse_data(link=None):
    try:

        # Устанавливаем путь к папке для скачивания
        download_dir = os.path.join(os.getcwd(), "downloads")

        s = Service(ChromeDriverManager().install())
        # Устанавливаем опции для веб-драйвера Chrome
        options = webdriver.ChromeOptions()
        options.add_argument("--window-size=1920,1080")
        options.add_experimental_option("prefs", {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": False,
            # "download.default_filename": "example.xls"
        })
        browser = webdriver.Chrome(service=s, options=options)

        # Загружаем страницу, на которой находится ссылка на скачивание файла
        browser.get(link or
                    "https://www.ups.com/us/en/support/shipping-support/shipping-costs-rates/retail-rates.page")

        # Находим инупт на вставку кода
        input_path = browser.find_element(By.XPATH, '//*[@id="ups-zone_zip"]')
        # return input_path

        # пробегаемся в цикле по нашим кодам и скачиваем
        zip_ranges = read_file()
        for zone_ranges in zip_ranges[:10]:
            # вставляем полученные коды из списка
            input_path.clear()

            zip_from_ranges = zone_ranges.split('-')[0]
            zip_to_ranges = zone_ranges.split('-')[1]

            input_path.send_keys(zip_from_ranges)

            button_to_download = \
                browser.find_element(By.XPATH,
                                     '//*[@id="iw_comp1631216335843"]'
                                     '/div/div/div/div/div/div[2]/form/div[2]/button')
            button_to_download.click()
            time.sleep(2)

            # открываем скаченный файл и проверяем на коды
            df = pd.read_excel(rf"./downloads/{zip_from_ranges[:3]}.xls")
            zip_from, zip_to = re.findall(r'\b(\d{3}-\d{2})\b',
                                          df.to_dict()['ZONE CHART'][3])
            del_defis = lambda x: int(x.replace('-', ''))
            del_defis_str = lambda x: x.replace('-', '')
            # print(zip_from)

            counter = 1
            # если коды не соответствуют то пробегаемся, качаем файлы и корректируем в списке
            while int(zip_from_ranges) + 1 != del_defis(zip_from) or \
                    int(zip_to_ranges) != del_defis(zip_to):

                new_zip_to_list = \
                    '0' + str(del_defis(zip_from) - 1) + '-' + del_defis_str(
                        zip_to)
                if counter == 1:
                    zip_ranges[zip_ranges.index(zone_ranges)] = new_zip_to_list

                new_zip_from = int(zip_from_ranges) + 100 * counter
                new_zip_to = new_zip_from + 99

                # если у нас есть в списке данный код мы пропускаем, иначе скачиваем и
                # добавляем код в лист
                if not any(('0' + str(new_zip_from) in i) for i in zip_ranges if
                           i):
                    input_path.clear()
                    input_path.send_keys('0' + str(new_zip_from))
                    button_to_download.click()
                    time.sleep(2)

                    df = pd.read_excel(
                        rf"./downloads/{'0' + str(new_zip_from)[:2]}.xls")
                    zip_from_, zip_to = re.findall(r'\b(\d{3}-\d{2})\b',
                                                   df.to_dict()['ZONE CHART'][
                                                       3])
                    range_to_insert = '0' + str(del_defis(zip_from_) - 1) \
                                      + '-' + del_defis_str(zip_to)

                    if range_to_insert not in zip_ranges:
                        zip_ranges.insert(
                            zip_ranges.index(new_zip_to_list) + counter,
                            range_to_insert)
                else:
                    zip_to = '0' + str(new_zip_to)
                counter += 1

        # Ждем завершения загрузки файла
        time.sleep(2)

        update_workbook(zip_ranges)


    except Exception as e:
        print(f'Exception occur: {e}')
    finally:
        browser.close()


# переименовываем файлы с правильным разрешением, так как они загружаются
# в формате xlsx, но с названием xls
def rename_files():
    folder_path = 'downloads'

    for filename in os.listdir(folder_path):
        if filename.endswith('.xls'):
            new_filename = filename[:-4] + '.xlsx'
            os.rename(os.path.join(folder_path, filename),
                      os.path.join(folder_path, new_filename))


if __name__ == '__main__':
    parse_data()
    rename_files()
    create_xlsx_alaska_hawaii()
