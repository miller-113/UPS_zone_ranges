import PyPDF2
import openpyxl
import re


def find_page(text_, reader):
    counter = 0

    while True:
        page = reader.pages[counter]
        text = page.extract_text()
        # if text.find('Determine the Zone 29'):
        # if text.find('Determine the Zone  11'):
        if text.find(text_) > 0:
            page_index = counter
            break
        counter += 1
    return page_index


def create_xlsx_alaska_hawaii(read_file_=None):
    # Открываем PDF-файл
    with open(read_file_ or 'ak-hi_retail_rates.pdf', 'rb') as pdf_file:
        # Создаем объект PyPDF2.PdfReader
        reader = PyPDF2.PdfReader(pdf_file)

        alaska_index = find_page('Determine the Zone  11', reader)
        hawaii_index = find_page('Determine the Zone  29', reader)

        # Получаем страницу документа
        page_alaska = reader.pages[alaska_index]
        page_hawaii = reader.pages[hawaii_index]

        # Извлекаем текст со страницы
        text_alaska = page_alaska.extract_text()
        text_hawaii = page_hawaii.extract_text()

        pattern_alaska = r'99\d{3}-?[0-9]*'
        pattern_hawaii = r'\d{5}-?\d{0,5}'

        # Ищем определенные данные в тексте
        data_alaska = re.findall(pattern_alaska, text_alaska)
        data_hawaii = re.findall(pattern_hawaii, text_hawaii)

        # создаем новый объект Workbook
        workbook = openpyxl.Workbook()

        # удаляем лист по умолчанию
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

        # создаем листы для Аляски и Гавайев
        sheet_alaska = workbook.create_sheet("Alaska")
        sheet_hawaii = workbook.create_sheet("Hawaii")

        # определяем хедер
        headers = ["UPS zone ranges", "Zip From", "Zip To"]

        # добавляем хедер в оба листа
        sheet_alaska.append(headers)
        sheet_hawaii.append(headers)

        # добавляем данные в лист Alaska
        for zip_range in data_alaska:
            if '-' in zip_range:
                zip_from, zip_to = zip_range.split("-")
                row = [zip_range, zip_from, zip_to]
                sheet_alaska.append(row)
            else:
                row = [zip_range]
                sheet_alaska.append(row)

        # добавляем данные в лист Hawaii
        for zip_range in data_hawaii:
            if '-' in zip_range:
                zip_from, zip_to = zip_range.split("-")
                row = [zip_range, zip_from, zip_to]
                sheet_hawaii.append(row)
            else:
                row = [zip_range]
                sheet_alaska.append(row)

        # сохраняем файл
        workbook.save("Carrierszoneranges_alaska_hawaii.xlsx")
        print('File was created!')
